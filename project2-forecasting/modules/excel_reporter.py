"""
excel_reporter.py
-----------------
Builds a 5-sheet Excel MRP report from PostgreSQL data.

Sheet 1: Executive Summary    — headline KPIs and waste totals
Sheet 2: Forecast Results     — 30-day Prophet forecast per SKU
Sheet 3: Reorder Parameters   — EOQ, ROP, safety stock, savings
Sheet 4: Lean Waste Flags     — all flagged SKUs with $ impact
Sheet 5: SAP Export           — MRP II flat file formatted for ERP import

Saves to: reports/mrp_report_YYYYMMDD.xlsx

Run standalone : python modules/excel_reporter.py
Called by      : main.py
"""

import os
import sys
import pandas as pd
from datetime import date, datetime
from pathlib import Path
from sqlalchemy import text

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from db.connection import get_engine


# ── Colour palette — consistent across all sheets ───────────────────────────
COLOUR = {
    "header_dark":  "1F3864",   # dark navy  — main headers
    "header_mid":   "2E75B6",   # mid blue   — sub headers
    "header_light": "D6E4F0",   # light blue — alternating rows
    "accent_green": "E2EFDA",   # light green
    "accent_red":   "FDECEA",   # light red
    "accent_amber": "FFF3CD",   # amber
    "white":        "FFFFFF",
    "text_white":   "FFFFFF",
    "text_dark":    "1F1F1F",
}


# ════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS — reused across all sheets
# ════════════════════════════════════════════════════════════════════════════

def style_header_row(ws, row_num, num_cols, dark=True):
    """Applies header styling to a full row."""
    fill_colour = COLOUR["header_dark"] if dark else COLOUR["header_mid"]
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = PatternFill("solid", fgColor=fill_colour)
        cell.font      = Font(bold=True, color=COLOUR["text_white"], size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def style_data_row(ws, row_num, num_cols, alternate=False):
    """Applies alternating row shading to data rows."""
    fill_colour = COLOUR["header_light"] if alternate else COLOUR["white"]
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill      = PatternFill("solid", fgColor=fill_colour)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_column_widths(ws, widths):
    """Sets column widths. widths = list of widths in order."""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def write_dataframe(ws, df, start_row, headers, col_widths,
                    dark_header=True, money_cols=None, float_cols=None):
    """
    Writes a DataFrame to a worksheet starting at start_row.
    Applies header styling, alternating row colours, and number formats.

    Parameters:
        headers    : list of display column names
        money_cols : list of 1-based column indices to format as $#,##0.00
        float_cols : list of 1-based column indices to format as #,##0.0
    """
    money_cols = money_cols or []
    float_cols = float_cols or []

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=header)
    style_header_row(ws, start_row, len(headers), dark=dark_header)

    # Write data rows
    for row_offset, (_, data_row) in enumerate(df.iterrows(), 1):
        row_num   = start_row + row_offset
        alternate = (row_offset % 2 == 0)

        for col_idx, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)

            # Apply money format
            if col_idx in money_cols:
                cell.number_format = '$#,##0.00'
            elif col_idx in float_cols:
                cell.number_format = '#,##0.0'

        style_data_row(ws, row_num, len(headers), alternate=alternate)

    set_column_widths(ws, col_widths)
    return start_row + len(df) + 1


# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Executive Summary
# Headline numbers only — designed for a 60-second read by a VP/Director
# ════════════════════════════════════════════════════════════════════════════
def build_sheet1_executive(ws, engine):
    """Builds the Executive Summary sheet."""

    ws.sheet_view.showGridLines = False

    # Pull headline numbers from PostgreSQL
    with engine.connect() as conn:
        total_skus = conn.execute(
            text("SELECT COUNT(DISTINCT sku) FROM reorder_params")
        ).scalar()

        at_risk = conn.execute(
            text("SELECT COUNT(*) FROM reorder_params "
                 "WHERE stockout_risk IN ('Critical','High')")
        ).scalar()

        total_waste = conn.execute(
            text("SELECT COALESCE(SUM(annual_waste_usd),0) "
                 "FROM lean_waste_flags")
        ).scalar()

        total_savings = conn.execute(
            text("SELECT COALESCE(SUM(potential_savings_usd),0) "
                 "FROM reorder_params")
        ).scalar()

        forecast_count = conn.execute(
            text("SELECT COUNT(*) FROM forecasts")
        ).scalar()

        risk_breakdown = pd.read_sql(text("""
            SELECT stockout_risk, COUNT(*) as sku_count
            FROM reorder_params
            GROUP BY stockout_risk
            ORDER BY
                CASE stockout_risk
                    WHEN 'Critical' THEN 1
                    WHEN 'High'     THEN 2
                    WHEN 'Medium'   THEN 3
                    WHEN 'Low'      THEN 4
                END
        """), conn)

        waste_breakdown = pd.read_sql(text("""
            SELECT waste_type,
                   COUNT(*) as flags,
                   SUM(annual_waste_usd) as total_waste_usd
            FROM lean_waste_flags
            GROUP BY waste_type
            ORDER BY total_waste_usd DESC
        """), conn)

    # ── Title block ──────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = "WAREHOUSE INVENTORY INTELLIGENCE — EXECUTIVE SUMMARY"
    title_cell.font      = Font(bold=True, size=16, color=COLOUR["text_white"])
    title_cell.fill      = PatternFill("solid", fgColor=COLOUR["header_dark"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:F2")
    sub_cell = ws["A2"]
    sub_cell.value     = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} "
                          f"  |  Pipeline: Project 2 — Forecasting & Lean Waste Engine")
    sub_cell.font      = Font(italic=True, size=10, color=COLOUR["text_white"])
    sub_cell.fill      = PatternFill("solid", fgColor=COLOUR["header_mid"])
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ── KPI Cards (row 4-5) ───────────────────────────────────────────────
    kpis = [
        ("Total SKUs Analysed",         total_skus,         None,     COLOUR["header_light"]),
        ("SKUs At Risk (Critical/High)", at_risk,            None,     COLOUR["accent_red"]),
        ("Forecast Rows Generated",      forecast_count,     None,     COLOUR["accent_green"]),
        ("Reorder Opt. Savings",         total_savings,      "$",      COLOUR["accent_green"]),
        ("Est. Annual Waste Cost",        total_waste,        "$",      COLOUR["accent_red"]),
    ]

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 30

    for col_idx, (label, value, prefix, colour) in enumerate(kpis, 1):
        # Label row
        label_cell = ws.cell(row=4, column=col_idx, value=label)
        label_cell.font      = Font(bold=True, size=9,
                                     color=COLOUR["text_dark"])
        label_cell.fill      = PatternFill("solid", fgColor=COLOUR["header_dark"])
        label_cell.font      = Font(bold=True, size=9, color="FFFFFF")
        label_cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Value row
        display_val = (f"${value:,.0f}" if prefix == "$"
                       else f"{value:,}")
        value_cell = ws.cell(row=5, column=col_idx, value=display_val)
        value_cell.font      = Font(bold=True, size=14,
                                     color=COLOUR["text_dark"])
        value_cell.fill      = PatternFill("solid", fgColor=colour)
        value_cell.alignment = Alignment(horizontal="center",
                                          vertical="center")

    # ── Stockout Risk Breakdown (row 8+) ─────────────────────────────────
    ws.cell(row=7, column=1).value = "STOCKOUT RISK BREAKDOWN"
    ws.cell(row=7, column=1).font  = Font(bold=True, size=11,
                                           color=COLOUR["header_dark"])

    risk_headers = ["Risk Tier", "SKU Count"]
    for col, h in enumerate(risk_headers, 1):
        ws.cell(row=8, column=col, value=h)
    style_header_row(ws, 8, 2)

    risk_colours = {
        "Critical": COLOUR["accent_red"],
        "High":     COLOUR["accent_red"],
        "Medium":   COLOUR["accent_amber"],
        "Low":      COLOUR["accent_green"],
    }

    for row_off, (_, r) in enumerate(risk_breakdown.iterrows(), 1):
        row_n = 8 + row_off
        ws.cell(row=row_n, column=1, value=r["stockout_risk"])
        ws.cell(row=row_n, column=2, value=int(r["sku_count"]))
        colour = risk_colours.get(r["stockout_risk"], COLOUR["white"])
        for col in range(1, 3):
            ws.cell(row=row_n, column=col).fill = PatternFill("solid",
                                                               fgColor=colour)
            ws.cell(row=row_n, column=col).alignment = Alignment(
                horizontal="center")

    # ── Waste Breakdown (row 8+ col 4+) ──────────────────────────────────
    ws.cell(row=7, column=4).value = "LEAN WASTE BREAKDOWN"
    ws.cell(row=7, column=4).font  = Font(bold=True, size=11,
                                           color=COLOUR["header_dark"])

    waste_headers = ["Waste Type", "Flags", "Annual Waste ($)"]
    for col, h in enumerate(waste_headers, 1):
        ws.cell(row=8, column=col+3, value=h)
    style_header_row(ws, 8, 3, dark=True)

    # Re-style columns 4-6 specifically for waste header
    for col in range(4, 7):
        ws.cell(row=8, column=col).fill = PatternFill("solid",
                                                        fgColor=COLOUR["header_dark"])
        ws.cell(row=8, column=col).font = Font(bold=True,
                                                color=COLOUR["text_white"])
        ws.cell(row=8, column=col).alignment = Alignment(horizontal="center")

    for row_off, (_, w) in enumerate(waste_breakdown.iterrows(), 1):
        row_n = 8 + row_off
        ws.cell(row=row_n, column=4, value=w["waste_type"])
        ws.cell(row=row_n, column=5, value=int(w["flags"]))
        ws.cell(row=row_n, column=6, value=round(float(w["total_waste_usd"]), 2))
        ws.cell(row=row_n, column=6).number_format = '$#,##0.00'
        for col in range(4, 7):
            ws.cell(row=row_n, column=col).fill = PatternFill(
                "solid", fgColor=COLOUR["accent_red"])
            ws.cell(row=row_n, column=col).alignment = Alignment(
                horizontal="center")

    set_column_widths(ws, [22, 14, 18, 28, 10, 18])
    print("   ✅ Sheet 1: Executive Summary")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Forecast Results
# ════════════════════════════════════════════════════════════════════════════
def build_sheet2_forecasts(ws, engine):
    """Builds the Forecast Results sheet."""

    ws.sheet_view.showGridLines = False

    with engine.connect() as conn:
        df = pd.read_sql(text("""
            SELECT
                sku,
                forecast_date,
                ROUND(yhat::numeric, 2)       AS predicted_orders,
                ROUND(yhat_lower::numeric, 2) AS lower_bound,
                ROUND(yhat_upper::numeric, 2) AS upper_bound,
                model_used,
                generated_at::date            AS generated_date
            FROM forecasts
            ORDER BY sku, forecast_date
        """), conn)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value     = "30-DAY DEMAND FORECAST RESULTS — All SKUs"
    ws["A1"].font      = Font(bold=True, size=13, color=COLOUR["text_white"])
    ws["A1"].fill      = PatternFill("solid", fgColor=COLOUR["header_dark"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers    = ["SKU", "Forecast Date", "Predicted Orders",
                  "Lower Bound", "Upper Bound", "Model", "Generated"]
    col_widths = [12, 16, 18, 14, 14, 18, 14]

    write_dataframe(
        ws, df, start_row=2,
        headers    = headers,
        col_widths = col_widths,
        money_cols = [],
        float_cols = [3, 4, 5],
    )

    print("   ✅ Sheet 2: Forecast Results")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Reorder Parameters
# ════════════════════════════════════════════════════════════════════════════
def build_sheet3_reorder(ws, engine):
    """Builds the Reorder Parameters sheet."""

    ws.sheet_view.showGridLines = False

    with engine.connect() as conn:
        df = pd.read_sql(text("""
            SELECT
                r.sku,
                s.category,
                ROUND(r.current_stock::numeric, 0)         AS current_stock,
                ROUND(r.days_of_stock::numeric, 1)         AS days_of_stock,
                r.lead_time_days,
                r.stockout_risk,
                ROUND(r.eoq::numeric, 1)                   AS eoq,
                ROUND(r.safety_stock::numeric, 1)          AS safety_stock,
                ROUND(r.current_rop::numeric, 1)           AS current_rop,
                ROUND(r.optimal_rop::numeric, 1)           AS optimal_rop,
                ROUND(r.gap::numeric, 1)                   AS gap,
                ROUND(r.potential_savings_usd::numeric, 2) AS potential_savings_usd
            FROM reorder_params r
            JOIN sku_master s USING (sku)
            ORDER BY r.potential_savings_usd DESC
        """), conn)

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"].value     = "REORDER PARAMETERS — EOQ, Safety Stock & ROP Optimisation"
    ws["A1"].font      = Font(bold=True, size=13, color=COLOUR["text_white"])
    ws["A1"].fill      = PatternFill("solid", fgColor=COLOUR["header_dark"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "SKU", "Category", "Current Stock", "Days of Stock",
        "Lead Time", "Risk", "EOQ", "Safety Stock",
        "Current ROP", "Optimal ROP", "Gap", "Potential Savings"
    ]
    col_widths = [12, 14, 14, 13, 11, 10, 8, 13, 13, 12, 8, 17]

    write_dataframe(
        ws, df, start_row=2,
        headers    = headers,
        col_widths = col_widths,
        money_cols = [12],
        float_cols = [3, 4, 7, 8, 9, 10, 11],
    )

    # Colour-code the Risk column (col 6) after writing
    risk_colours = {
        "Critical": COLOUR["accent_red"],
        "High":     COLOUR["accent_red"],
        "Medium":   COLOUR["accent_amber"],
        "Low":      COLOUR["accent_green"],
    }
    for row_idx in range(3, 3 + len(df)):
        risk_val = ws.cell(row=row_idx, column=6).value
        colour   = risk_colours.get(risk_val, COLOUR["white"])
        ws.cell(row=row_idx, column=6).fill = PatternFill("solid",
                                                           fgColor=colour)

    print("   ✅ Sheet 3: Reorder Parameters")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Lean Waste Flags
# ════════════════════════════════════════════════════════════════════════════
def build_sheet4_waste(ws, engine):
    """Builds the Lean Waste Flags sheet."""

    ws.sheet_view.showGridLines = False

    with engine.connect() as conn:
        df = pd.read_sql(text("""
            SELECT
                sku,
                waste_type,
                severity,
                ROUND(annual_waste_usd::numeric, 2) AS annual_waste_usd,
                detail,
                flagged_at::date                    AS flagged_date
            FROM lean_waste_flags
            ORDER BY annual_waste_usd DESC
        """), conn)

        total_waste = df["annual_waste_usd"].sum()

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"].value     = "LEAN WASTE FLAGS — Excess Inventory, Over-Ordering & Demand Planning"
    ws["A1"].font      = Font(bold=True, size=13, color=COLOUR["text_white"])
    ws["A1"].fill      = PatternFill("solid", fgColor=COLOUR["header_dark"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Totals row before data
    ws.merge_cells("A2:C2")
    ws["A2"].value     = (f"Total Estimated Annual Waste: "
                          f"${total_waste:,.2f} across "
                          f"{df['sku'].nunique()} SKUs "
                          f"({len(df)} flags)")
    ws["A2"].font      = Font(bold=True, size=11, color=COLOUR["text_white"])
    ws["A2"].fill      = PatternFill("solid", fgColor=COLOUR["header_mid"])
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    headers    = ["SKU", "Waste Type", "Severity",
                  "Annual Waste ($)", "Detail", "Flagged Date"]
    col_widths = [12, 26, 10, 18, 60, 14]

    write_dataframe(
        ws, df, start_row=3,
        headers    = headers,
        col_widths = col_widths,
        money_cols = [4],
    )

    # Colour-code severity (col 3) after writing
    sev_colours = {
        "High":   COLOUR["accent_red"],
        "Medium": COLOUR["accent_amber"],
        "Low":    COLOUR["accent_green"],
    }
    for row_idx in range(4, 4 + len(df)):
        sev_val = ws.cell(row=row_idx, column=3).value
        colour  = sev_colours.get(sev_val, COLOUR["white"])
        ws.cell(row=row_idx, column=3).fill = PatternFill("solid",
                                                           fgColor=colour)
        # Wrap the detail column text
        ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True,
                                                              vertical="top")
        ws.row_dimensions[row_idx].height = 40

    print("   ✅ Sheet 4: Lean Waste Flags")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — SAP Export
# ════════════════════════════════════════════════════════════════════════════
def build_sheet5_sap(ws, engine):
    """Builds the SAP Export sheet — MRP II flat file format."""

    ws.sheet_view.showGridLines = False

    with engine.connect() as conn:
        df = pd.read_sql(text("""
            SELECT
                sku                                    AS MATNR,
                'WH01'                                 AS WERKS,
                GREATEST(1, ROUND(optimal_rop::numeric, 0)::int)        AS MINBE,
                GREATEST(0, ROUND(safety_stock::numeric, 0)::int)       AS EISBE,
                GREATEST(1, ROUND(optimal_rop::numeric*3, 0)::int)      AS MABST,
                GREATEST(1, ROUND(eoq::numeric*0.5, 0)::int)            AS BSTMI,
                GREATEST(1, ROUND(eoq::numeric*2, 0)::int)              AS BSTMA
            FROM reorder_params
            ORDER BY sku
        """), conn)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value     = "SAP MM MRP II EXPORT — Transaction MM17 Import Ready"
    ws["A1"].font      = Font(bold=True, size=13, color=COLOUR["text_white"])
    ws["A1"].fill      = PatternFill("solid", fgColor=COLOUR["header_dark"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Field reference row
    ws.merge_cells("A2:G2")
    ws["A2"].value = ("MATNR=Material | WERKS=Plant | MINBE=Reorder Pt | "
                      "EISBE=Safety Stock | MABST=Max Stock | "
                      "BSTMI=Min Lot Size | BSTMA=Max Lot Size")
    ws["A2"].font      = Font(italic=True, size=9,
                               color=COLOUR["text_white"])
    ws["A2"].fill      = PatternFill("solid", fgColor=COLOUR["header_mid"])
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    headers    = ["MATNR", "WERKS", "MINBE", "EISBE", "MABST", "BSTMI", "BSTMA"]
    col_widths = [14, 10, 10, 10, 12, 12, 12]

    write_dataframe(
        ws, df, start_row=3,
        headers    = headers,
        col_widths = col_widths,
    )

    print("   ✅ Sheet 5: SAP Export")


# ════════════════════════════════════════════════════════════════════════════
# MAIN — assembles all 5 sheets into one workbook
# ════════════════════════════════════════════════════════════════════════════
def run_excel_reporter(engine=None):
    """
    Builds the full 5-sheet MRP Excel report.
    Returns the output file path.
    """
    print("\n" + "=" * 55)
    print("  Module 6 — Excel MRP Report Builder")
    print("=" * 55)

    if engine is None:
        engine = get_engine()

    # Create workbook and name the sheets
    wb = openpyxl.Workbook()

    # openpyxl creates one default sheet — rename it for Sheet 1
    wb.active.title = "Executive Summary"
    ws1 = wb.active

    ws2 = wb.create_sheet("Forecast Results")
    ws3 = wb.create_sheet("Reorder Parameters")
    ws4 = wb.create_sheet("Lean Waste Flags")
    ws5 = wb.create_sheet("SAP Export")

    print("\n   Building sheets...")
    build_sheet1_executive(ws1, engine)
    build_sheet2_forecasts(ws2, engine)
    build_sheet3_reorder(ws3, engine)
    build_sheet4_waste(ws4, engine)
    build_sheet5_sap(ws5, engine)

    # Save to reports/ folder
    reports_dir = Path(__file__).parent.parent / "reports"
    reports_dir.mkdir(exist_ok=True)

    filename  = f"mrp_report_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath  = reports_dir / filename
    wb.save(filepath)

    print(f"\n   💾 MRP Report saved:")
    print(f"      {filepath}")
    print(f"\n   ✅ Excel Reporter complete!")
    print("=" * 55)

    return str(filepath)


if __name__ == "__main__":
    run_excel_reporter()